"""
Extractor de Comprobantes Peruanos con IA
==========================================
Procesa lotes de facturas, boletas y recibos en PDF y genera un Excel
estructurado con validación SUNAT y formato profesional.

Uso:
    python extractor.py                          # usa carpeta 'pdfs/' por defecto
    python extractor.py mis_facturas             # carpeta custom
    python extractor.py mis_facturas reporte.xlsx
"""

import pdfplumber
import google.generativeai as genai
import pandas as pd
from pathlib import Path
import json
import os
import re
import sys
from datetime import datetime
from tqdm import tqdm
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN
# ============================================================
CARPETA_PDFS_DEFAULT = "pdfs"
ARCHIVO_SALIDA_DEFAULT = "output/facturas_extraidas.xlsx"
ARCHIVO_LOG = "output/log_procesamiento.txt"
MODELO_IA = "gemini-flash-latest"


# ============================================================
# VALIDACIONES SUNAT (PERÚ)
# ============================================================
def validar_ruc(ruc) -> bool:
    """Valida un RUC peruano con el algoritmo oficial del dígito verificador."""
    if not ruc:
        return False
    ruc = str(ruc).strip()
    if not ruc.isdigit() or len(ruc) != 11:
        return False
    factores = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    suma = sum(int(ruc[i]) * factores[i] for i in range(10))
    resto = suma % 11
    digito_calculado = (11 - resto) % 10
    return digito_calculado == int(ruc[10])


def validar_formato_serie(numero_documento) -> bool:
    """Valida formato de comprobante peruano: F001-XXXXX, B001-XXXXX, etc."""
    if not numero_documento:
        return False
    limpio = str(numero_documento).replace(" ", "").upper()
    return bool(re.match(r"^[FB][A-Z0-9]{3}-\d+$", limpio))


# ============================================================
# CONFIGURACIÓN DE LA IA
# ============================================================
def configurar_ia():
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("❌ No se encontró GEMINI_API_KEY en las variables de entorno.")
        print('   Ejecuta: $env:GEMINI_API_KEY="tu_clave"')
        sys.exit(1)
    genai.configure(api_key=api_key)
    return genai.GenerativeModel(MODELO_IA)


# ============================================================
# EXTRACCIÓN DE TEXTO Y DATOS
# ============================================================
def extraer_texto(ruta_pdf: str) -> str:
    """Extrae todo el texto de un PDF."""
    texto_completo = []
    with pdfplumber.open(ruta_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""
            texto_completo.append(texto)
    return "\n".join(texto_completo)


def extraer_datos_factura(texto_factura: str, modelo) -> dict:
    """Usa IA para extraer datos estructurados de un comprobante."""
    prompt = f"""Eres un experto en comprobantes peruanos (facturas, boletas, recibos por honorarios, notas de crédito).

Analiza el siguiente texto y devuelve EXCLUSIVAMENTE un JSON válido con esta estructura, sin markdown ni explicaciones:

{{
  "tipo_documento": "factura | boleta | recibo_honorarios | nota_credito | otro",
  "ruc_emisor": "RUC del emisor (11 dígitos) o null",
  "razon_social_emisor": "Nombre o razón social del emisor o null",
  "ruc_cliente": "RUC/DNI del cliente o null",
  "razon_social_cliente": "Nombre del cliente o null",
  "numero_documento": "Serie y número (ej: F001-00123) o null",
  "fecha_emision": "YYYY-MM-DD o null",
  "moneda": "PEN | USD | otro",
  "subtotal": número decimal o null,
  "igv": número decimal o null,
  "total": número decimal o null,
  "items": [
    {{"descripcion": "...", "cantidad": número, "precio_unitario": número, "subtotal_item": número}}
  ]
}}

Reglas:
- Si un campo no está presente, usa null
- Para números, devuelve solo el valor numérico (sin S/, $, comas)
- Si el documento NO es un comprobante de pago, igual extrae lo que puedas y marca tipo_documento como "otro"

Texto del comprobante:
\"\"\"
{texto_factura}
\"\"\"
"""
    response = modelo.generate_content(prompt)
    texto_respuesta = response.text.strip()
    texto_respuesta = re.sub(r"^```(?:json)?\s*", "", texto_respuesta)
    texto_respuesta = re.sub(r"\s*```$", "", texto_respuesta)
    return json.loads(texto_respuesta)


# ============================================================
# FORMATO PROFESIONAL DEL EXCEL
# ============================================================
def aplicar_formato_excel(writer, hojas: dict):
    """Aplica formato profesional: encabezados con color, anchos auto, moneda."""
    workbook = writer.book

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    cols_moneda = {"subtotal", "igv", "total", "precio_unitario", "subtotal_item"}

    for sheet_name, df in hojas.items():
        if df is None or df.empty:
            continue
        ws = workbook[sheet_name]

        # Encabezados
        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        # Anchos automáticos (max 50)
        for col_idx, col in enumerate(df.columns, start=1):
            letra = get_column_letter(col_idx)
            valores = [str(v) for v in df[col].astype(str).tolist() if v not in ("nan", "None")]
            max_len = max([len(str(col))] + [len(v) for v in valores] + [10])
            ws.column_dimensions[letra].width = min(max_len + 3, 50)

        # Altura encabezado y congelar
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Bordes y formato moneda en datos
        for col_idx, col in enumerate(df.columns, start=1):
            letra = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{letra}{row}"]
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if col in cols_moneda:
                    cell.number_format = "#,##0.00"


# ============================================================
# PROCESAMIENTO PRINCIPAL
# ============================================================
def procesar_carpeta(carpeta_pdfs: str, archivo_salida: str):
    carpeta = Path(carpeta_pdfs)
    if not carpeta.exists():
        print(f"❌ La carpeta '{carpeta}' no existe.")
        return

    pdfs = sorted(carpeta.glob("*.pdf"))
    if not pdfs:
        print(f"❌ No se encontraron PDFs en '{carpeta}'.")
        return

    Path(archivo_salida).parent.mkdir(parents=True, exist_ok=True)

    # Verificar que el Excel no esté abierto
    salida_path = Path(archivo_salida)
    if salida_path.exists():
        try:
            with open(salida_path, "a"):
                pass
        except PermissionError:
            print(f"❌ El archivo '{archivo_salida}' está abierto en Excel.")
            print("   Ciérralo y vuelve a ejecutar.")
            return

    print(f"\n{'='*60}")
    print(f"  📄 EXTRACTOR DE COMPROBANTES — {len(pdfs)} archivo(s)")
    print(f"{'='*60}\n")

    modelo = configurar_ia()
    resultados = []
    items_detalle = []
    log_registros = []
    inicio = datetime.now()

    for pdf_path in tqdm(pdfs, desc="Procesando", unit="pdf", ncols=70):
        registro = {"archivo": pdf_path.name, "estado": "ok", "error": ""}
        try:
            texto = extraer_texto(str(pdf_path))
            if not texto.strip():
                raise ValueError("PDF sin texto extraíble (posiblemente escaneado, requiere OCR)")

            datos = extraer_datos_factura(texto, modelo)
            datos["archivo"] = pdf_path.name

            # Validaciones SUNAT
            datos["ruc_emisor_valido"] = "✓" if validar_ruc(datos.get("ruc_emisor")) else "✗"
            datos["ruc_cliente_valido"] = "✓" if validar_ruc(datos.get("ruc_cliente")) else "✗"
            datos["serie_valida"] = "✓" if validar_formato_serie(datos.get("numero_documento")) else "✗"

            resumen = {k: v for k, v in datos.items() if k != "items"}
            resultados.append(resumen)

            for item in datos.get("items") or []:
                item["archivo"] = pdf_path.name
                item["numero_documento"] = datos.get("numero_documento")
                items_detalle.append(item)

        except Exception as e:
            registro["estado"] = "error"
            registro["error"] = str(e)[:300]
            resultados.append({"archivo": pdf_path.name, "error": str(e)[:300]})

        log_registros.append(registro)

    duracion = (datetime.now() - inicio).total_seconds()

    # Generar Excel formateado
    df_resumen = pd.DataFrame(resultados)
    df_items = pd.DataFrame(items_detalle) if items_detalle else None

    with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        if df_items is not None and not df_items.empty:
            df_items.to_excel(writer, sheet_name="Detalle_Items", index=False)
        hojas = {"Resumen": df_resumen, "Detalle_Items": df_items}
        aplicar_formato_excel(writer, hojas)

    # Log de procesamiento
    with open(ARCHIVO_LOG, "w", encoding="utf-8") as f:
        f.write(f"Procesamiento del {inicio.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Duración: {duracion:.1f} segundos\n")
        f.write(f"Total archivos: {len(pdfs)}\n")
        exitosos = sum(1 for r in log_registros if r["estado"] == "ok")
        f.write(f"Exitosos: {exitosos}\n")
        f.write(f"Con error: {len(log_registros) - exitosos}\n\n")
        for r in log_registros:
            f.write(f"[{r['estado'].upper()}] {r['archivo']}")
            if r["error"]:
                f.write(f" — {r['error']}")
            f.write("\n")

    # Resumen final en consola
    exitosos = sum(1 for r in log_registros if r["estado"] == "ok")
    fallidos = len(log_registros) - exitosos

    total_pen = sum(
        r.get("total", 0) or 0 for r in resultados
        if r.get("moneda") == "PEN" and isinstance(r.get("total"), (int, float))
    )
    total_usd = sum(
        r.get("total", 0) or 0 for r in resultados
        if r.get("moneda") == "USD" and isinstance(r.get("total"), (int, float))
    )

    print(f"\n{'='*60}")
    print(f"  ✅ PROCESAMIENTO COMPLETADO")
    print(f"{'='*60}")
    print(f"  Archivos procesados:  {len(pdfs)}")
    print(f"  Exitosos:             {exitosos}")
    print(f"  Con error:            {fallidos}")
    print(f"  Duración:             {duracion:.1f} segundos")
    if total_pen > 0:
        print(f"  Total facturado PEN:  S/ {total_pen:,.2f}")
    if total_usd > 0:
        print(f"  Total facturado USD:  $ {total_usd:,.2f}")
    print(f"\n  📊 Excel:  {archivo_salida}")
    print(f"  📝 Log:    {ARCHIVO_LOG}")
    print(f"{'='*60}\n")


# ============================================================
# PUNTO DE ENTRADA
# ============================================================
if __name__ == "__main__":
    carpeta = sys.argv[1] if len(sys.argv) > 1 else CARPETA_PDFS_DEFAULT
    salida = sys.argv[2] if len(sys.argv) > 2 else ARCHIVO_SALIDA_DEFAULT
    procesar_carpeta(carpeta, salida)